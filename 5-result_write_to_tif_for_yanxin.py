# -*- coding: utf-8 -*-
"""
预测表格文件转化为tif
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
Created on Thu Jul 29 11:25:34 2021
Reconstructed by wHy 2023.01.01
"""

import numpy as np
from skimage import io
from openpyxl import load_workbook
import sys


if __name__ == '__main__':

    sr_data_path = r'E:\yanxin\predict\predict_test_100.xlsx'  # 原始预测结果路径
    creatpath = r'E:\yanxin\predict\result_test_100_230103.tif'  # 转换tif路径
    row = 1240  # 输出影像行数
    col = 320  # 输出影像列数

    if len(sys.argv) > 2:
        sr_data_path = sys.argv[1]
        creatpath = sys.argv[2]
        row = int(sys.argv[3])
        col = int(sys.argv[4])

    print('读取数据...')
    book = load_workbook(filename=sr_data_path)
    print('读取数据完毕')
    print('-----------------------------------------------------')
    sheet = book.get_sheet_by_name("pred")
    data = []
    row1 = sheet.max_row + 1  # 行
    col1 = sheet.max_column + 1  # 列
    
    print('表格转图像...')
    for i in range(2, row1):
        for j in range(2, col1):
            data.append(sheet.cell(row=i, column=j).value)

    middle = np.zeros([row, col])

    for i in range(0, row):
        for j in range(0, col):
            middle[i][j] = data[i * col + j]

    io.imsave(creatpath, middle)
    print('表格换图像完毕')
